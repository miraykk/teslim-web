import io
import re
import zipfile
from pathlib import Path
from typing import Optional, List, Tuple
from dataclasses import dataclass, field
from datetime import date

import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pypdf import PdfReader, PdfWriter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# ── Sayfa yapılandırması ────────────────────────────────────────────────────
st.set_page_config(
    page_title="Teslim Tesellüm Oluşturucu",
    page_icon="📄",
    layout="wide",
)

# ── Sabit listeler ──────────────────────────────────────────────────────────
FIRMA_LISTESI = ["akakçe", "cevher", "çıtır", "altınsa"]
ISLEM_LISTESI = ["Alış", "Satış"]

DEFAULT_COL_MAP = {
    "TARIH_COL": 1,
    "FATURA_COL": 2,
    "ISIM_COL": 3,
    "ADRES_COL": 4,
    "TC1_COL": 5,
    "TC2_COL": 6,
    "GR_COL": 8,
    "TL_COL": 10,
}

TEMPLATES_DIR = Path(__file__).parent / "templates"

# ── Yardımcı fonksiyonlar ───────────────────────────────────────────────────

def normalize_tr(s: str) -> str:
    s = (s or "").lower().strip()
    return (s.replace("ı", "i").replace("ç", "c").replace("ğ", "g")
              .replace("ş", "s").replace("ö", "o").replace("ü", "u"))


def safe_filename(name: str) -> str:
    name = str(name).strip()
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    name = re.sub(r"\s+", "", name)
    return name or "bos_fatura"


def first_non_empty(ws, row: int, cols: List[int]):
    for c in cols:
        v = ws.cell(row, c).value
        if v is not None and str(v).strip() != "":
            return v
    return None


def kisa_firma_adi(name: str) -> str:
    if name is None:
        return ""
    s = str(name).strip().upper()
    replacements = {
        "LİMİTED": "LTD.", "LIMITED": "LTD.", "ŞİRKETİ": "ŞTİ.",
        "ANONİM": "A.Ş.", "SANAYİ": "SAN.", "TİCARET": "TİC.",
        "PAZARLAMA": "PAZ.", "HİZMETLERİ": "HİZ.", "İNŞAAT": "İNŞ.",
        "NAKLİYE": "NAK.", "LOJİSTİK": "LOJ.", "TEKNOLOJİ": "TEK.",
        "MÜHENDİSLİK": "MÜH.", "DANIŞMANLIK": "DAN.", "MUHASEBE": "MUH.",
    }
    for old, new in sorted(replacements.items(), key=lambda x: len(x[0]), reverse=True):
        s = re.sub(rf"\b{re.escape(old)}\b", new, s)
    return re.sub(r"\s+", " ", s).strip()


def template_path(firma: str, islem: str) -> Path:
    f = normalize_tr(firma)
    t = "alis" if islem.lower().startswith("al") else "satis"
    return TEMPLATES_DIR / f"{f}_{t}.xlsx"


def format_date(val) -> str:
    if val is None:
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%d.%m.%Y")
    return str(val).strip()


def format_number(val, decimals=2) -> str:
    if val is None:
        return ""
    try:
        return f"{float(val):,.{decimals}f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(val)


# ── PDF üretimi (reportlab) ─────────────────────────────────────────────────

def build_pdf_bytes(
    firma: str,
    islem_turu: str,
    tarih, fatura, isim, adres, tc_or_vkn, gr, tl,
) -> bytes:
    """Tek bir teslim-tesellüm belgesi PDF'i üretir, bytes olarak döner."""

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    bold_style = ParagraphStyle("bold", parent=styles["Normal"],
                                 fontName="Helvetica-Bold", fontSize=11, alignment=TA_CENTER)
    title_style = ParagraphStyle("title", parent=styles["Normal"],
                                  fontName="Helvetica-Bold", fontSize=13, alignment=TA_CENTER,
                                  spaceAfter=4)
    small_style = ParagraphStyle("small", parent=styles["Normal"],
                                  fontName="Helvetica", fontSize=8, alignment=TA_LEFT,
                                  leading=10)

    # Firma ve belge başlığı
    f_norm = normalize_tr(firma)
    firma_display = firma.upper()
    islem_display = "ALIMINA" if islem_turu == "Alış" else "SATIMINA"

    kisa_isim = kisa_firma_adi(isim) if isim else ""

    alici_label = "Alıcı Firma" if islem_turu == "Alış" else "Satıcı Müşteri / Firma"
    satici_label = "Satıcı Müşteri / Firma" if islem_turu == "Alış" else "Alıcı Firma"

    firma_unvan = "Akakçe Kıymetli Madenler San. ve Tic. Ltd. Şti."
    vergi_no = "0111231210"

    is_akakce = f_norm == "akakce"

    # Teslim yeri
    teslim_yeri = "Fatih / İstanbul" if is_akakce else ""

    tablo_data = [
        ["KONU - CİNS", "kıymetli madenler"],
        ["TESLİM TARİHİ", format_date(tarih)],
        ["TESLİM YERİ", teslim_yeri],
        ["FATURA NO", str(fatura) if fatura else ""],
        ["MİKTAR (gr)", format_number(gr, 4)],
        ["TOPLAM TUTAR (TL)", format_number(tl, 2)],
    ]

    firma_tablo = [
        [alici_label, ""],
        ["UNVAN", firma_unvan],
        ["VERGİ NO", vergi_no],
        [satici_label, ""],
        ["AD SOYAD / UNVAN", kisa_isim],
        ["T.C. / VERGİ NO", str(tc_or_vkn) if tc_or_vkn else ""],
        ["ADRES", str(adres) if adres else ""],
    ]

    disclaimer = (
        "Satın alınan kıymetli maden, Kıymetli Maden Alım Satım Belgesi ve ilgili fatura "
        "kapsamında, belirtilen miktarda tam, eksiksiz ve ayıpsız olarak teslim alınmış olup "
        "bu belge taraflarca tanzim edilmiştir.\n\n"
        "Müşteri Tanı ve İşlem Formu vasıtasıyla elde edilen kişisel verilerimin, 5549 sayılı "
        "Kanun'un 3. maddesi ve 5337 sayılı Yönetmelik 5. ve 6. maddeleri uyarınca kimlik "
        "tespiti ve müşteri tanıma yükümlülükleri kapsamında işlenmesine açık rızam;\n"
        "☐  vardır          ☐  yoktur"
    )

    imza_data = [
        ["ALTINI TESLİM EDEN", "ALTINI TESLİM ALAN"],
        [kisa_isim if islem_turu == "Alış" else firma_unvan,
         firma_unvan if islem_turu == "Alış" else kisa_isim],
    ]

    col_w = [5*cm, 11.5*cm]
    col_w2 = [8*cm, 8.5*cm]

    ts_main = TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f0f0")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("ROWBACKGROUND", (0, 0), (-1, -1), [colors.white, colors.HexColor("#fafafa")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ])

    ts_firma = TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f0f0f0")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("SPAN", (0, 0), (1, 0)),
        ("SPAN", (0, 3), (1, 3)),
        ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#d0e4f7")),
        ("BACKGROUND", (0, 3), (1, 3), colors.HexColor("#d0e4f7")),
        ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 3), (1, 3), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (1, 0), 9),
        ("FONTSIZE", (0, 3), (1, 3), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ])

    ts_imza = TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 14),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
    ])

    story = [
        Paragraph(f"{firma_display} KIYMETLİ MADENLER", title_style),
        Paragraph(f"KIYMETLİ MADEN {islem_display} İLİŞKİN", bold_style),
        Paragraph("TESLİM TESELLÜM BELGESİ", bold_style),
        Spacer(1, 0.4*cm),
        Table(tablo_data, colWidths=col_w, style=ts_main),
        Spacer(1, 0.3*cm),
        Table(firma_tablo, colWidths=col_w, style=ts_firma),
        Spacer(1, 0.3*cm),
        Paragraph(disclaimer.replace("\n", "<br/>"), small_style),
        Spacer(1, 0.5*cm),
        Table(imza_data, colWidths=col_w2, style=ts_imza),
    ]

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ── İş mantığı ─────────────────────────────────────────────────────────────

def load_input_xlsx(file_bytes: bytes) -> Tuple[object, object]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    return wb, ws


def get_headers(file_bytes: bytes) -> List[str]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        v = cell.value
        headers.append("" if v is None else str(v).strip())
    wb.close()
    return headers


def header_to_col_index(headers: List[str], chosen: str) -> Optional[int]:
    chosen = (chosen or "").strip()
    if not chosen or chosen == "Boş bırak":
        return None
    for i, h in enumerate(headers, 1):
        if (h or "").strip() == chosen:
            return i
    return None


def run_job(
    firma: str,
    islem_turu: str,
    input_bytes: bytes,
    col_map: dict,
    group_mode: str,
    progress_bar,
    status_text,
) -> bytes:
    """Tüm işlemi yapar, sonucu ZIP bytes olarak döner."""

    wb, ws = load_input_xlsx(input_bytes)

    TARIH_COL  = col_map.get("TARIH_COL")
    FATURA_COL = col_map.get("FATURA_COL")
    ISIM_COL   = col_map.get("ISIM_COL")
    ADRES_COL  = col_map.get("ADRES_COL")
    TC1_COL    = col_map.get("TC1_COL")
    TC2_COL    = col_map.get("TC2_COL")
    GR_COL     = col_map.get("GR_COL")
    TL_COL     = col_map.get("TL_COL")
    TC_COLS    = [c for c in [TC1_COL, TC2_COL] if c]

    # Geçerli satırları topla
    valid_rows = []
    for r in range(2, ws.max_row + 1):
        if FATURA_COL:
            fatura = ws.cell(r, FATURA_COL).value
            if fatura is None or str(fatura).strip() == "":
                continue
        valid_rows.append(r)

    if not valid_rows:
        raise ValueError("Dosyada işlenecek satır bulunamadı (Fatura No kolonunu kontrol edin).")

    total = len(valid_rows)
    pdf_files: dict = {}  # filename -> bytes

    for i, r in enumerate(valid_rows):
        tarih     = ws.cell(r, TARIH_COL).value if TARIH_COL else None
        fatura    = ws.cell(r, FATURA_COL).value if FATURA_COL else f"satir_{r}"
        isim      = ws.cell(r, ISIM_COL).value if ISIM_COL else None
        adres     = ws.cell(r, ADRES_COL).value if ADRES_COL else None
        tc_or_vkn = first_non_empty(ws, r, TC_COLS) if TC_COLS else None
        gr        = ws.cell(r, GR_COL).value if GR_COL else None
        tl        = ws.cell(r, TL_COL).value if TL_COL else None

        fname = safe_filename(fatura)
        pdf_bytes = build_pdf_bytes(firma, islem_turu, tarih, fatura, isim, adres, tc_or_vkn, gr, tl)
        pdf_files[fname] = (tarih, str(fatura), tc_or_vkn, pdf_bytes)

        pct = int((i + 1) / total * 90)
        progress_bar.progress(pct)
        status_text.text(f"İşleniyor: {i+1}/{total} — {fname}")

    status_text.text("PDF'ler birleştiriliyor…")
    progress_bar.progress(95)

    # ZIP oluştur
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Ayrı PDF'leri ekle
        for fname, (tarih, fatura_no, tc_or_vkn, pdf_b) in pdf_files.items():
            zf.writestr(f"pdfler/{fname}.pdf", pdf_b)

        if group_mode == "Tüm PDF'ler ortak":
            # Hepsini tarih+fatura sırasına göre birleştir
            items_sorted = sorted(
                pdf_files.items(),
                key=lambda x: (
                    x[1][0].isoformat() if hasattr(x[1][0], "isoformat") else str(x[1][0] or ""),
                    x[1][1]
                )
            )
            merged = _merge_pdf_list([v[3] for _, v in items_sorted])
            zf.writestr("tum_pdfler.pdf", merged)

        else:  # TC/VKN'ye göre gruplu
            groups: dict = {}
            for fname, (tarih, fatura_no, tc_or_vkn, pdf_b) in pdf_files.items():
                gkey = safe_filename(str(tc_or_vkn)) if tc_or_vkn else "NO_TC_VKN"
                groups.setdefault(gkey, []).append(
                    (tarih.isoformat() if hasattr(tarih, "isoformat") else str(tarih or ""),
                     fatura_no, pdf_b)
                )
            for gkey, items in groups.items():
                sorted_items = sorted(items, key=lambda x: (x[0], x[1]))
                merged = _merge_pdf_list([it[2] for it in sorted_items])
                zf.writestr(f"gruplu_pdfler/{gkey}.pdf", merged)

    progress_bar.progress(100)
    status_text.text(f"✅ Tamamlandı! {total} belge işlendi.")
    zip_buf.seek(0)
    return zip_buf.read()


def _merge_pdf_list(pdf_bytes_list: List[bytes]) -> bytes:
    writer = PdfWriter()
    for pdf_b in pdf_bytes_list:
        reader = PdfReader(io.BytesIO(pdf_b))
        for page in reader.pages:
            writer.add_page(page)
    buf = io.BytesIO()
    writer.write(buf)
    buf.seek(0)
    return buf.read()


# ── Streamlit Arayüzü ───────────────────────────────────────────────────────

def main():
    st.title("📄 Teslim Tesellüm Oluşturucu")
    st.caption("Excel listesinden otomatik teslim-tesellüm belgesi üretir.")

    st.divider()

    # ─ Ayarlar ─
    col1, col2, col3 = st.columns([2, 2, 3])
    with col1:
        firma = st.selectbox("Firma", FIRMA_LISTESI)
    with col2:
        islem_turu = st.selectbox("İşlem Türü", ISLEM_LISTESI)
    with col3:
        group_mode = st.selectbox(
            "PDF Birleştirme",
            ["Tüm PDF'ler ortak", "TC / Vergi No'ya göre gruplu"],
        )

    st.divider()

    # ─ Dosya yükleme ─
    uploaded = st.file_uploader(
        "📂 Esnek Rapor (.xlsx) yükle",
        type=None,
        help="İlk satır başlık olmalı. Fatura No kolonu boş olan satırlar atlanır.",
    )

    headers = []
    file_bytes = None

    if uploaded is not None:
        file_bytes = uploaded.read()
        try:
            headers = get_headers(file_bytes)
            st.success(f"✔ Dosya okundu — {len(headers)} kolon bulundu.")
        except Exception as e:
            st.error(f"Dosya okunamadı: {e}")
            return

    st.divider()

    # ─ Kolon eşleme ─
    with st.expander("⚙️ Kolon Eşleme (Opsiyonel)", expanded=bool(headers)):
        st.caption("Boş bırakılan kolonlar varsayılan sıraya göre atanır.")

        header_opts = ["Boş bırak"] + [h for h in headers if h.strip()]

        fields = [
            ("Tarih", "TARIH_COL", 1),
            ("Fatura No", "FATURA_COL", 2),
            ("İsim / Unvan", "ISIM_COL", 3),
            ("Adres", "ADRES_COL", 4),
            ("TC Kimlik No", "TC1_COL", 5),
            ("Vergi No (varsa)", "TC2_COL", 6),
            ("Gram", "GR_COL", 8),
            ("Tutar (TL)", "TL_COL", 10),
        ]

        col_selections = {}
        col_a, col_b = st.columns(2)
        for idx, (label, key, default_col) in enumerate(fields):
            container = col_a if idx % 2 == 0 else col_b
            with container:
                if headers and 1 <= default_col <= len(headers) and headers[default_col - 1].strip():
                    default_val = headers[default_col - 1]
                    default_idx = header_opts.index(default_val) if default_val in header_opts else 0
                else:
                    default_idx = 0
                sel = st.selectbox(
                    label,
                    options=header_opts,
                    index=default_idx,
                    key=f"col_{key}",
                    disabled=not headers,
                )
                col_selections[key] = sel

    # Kolon indekslerini çöz
    col_map = {}
    for key, chosen in col_selections.items():
        col_map[key] = header_to_col_index(headers, chosen)

    # Eğer hiç eşleme yapılmadıysa varsayılanı kullan
    any_selected = any(v is not None for v in col_map.values())
    if not any_selected:
        col_map = dict(DEFAULT_COL_MAP)

    st.divider()

    # ─ Çalıştır ─
    tpl = template_path(firma, islem_turu)
    if not tpl.exists():
        st.warning(
            f"⚠️ **{firma}** firması için **{islem_turu}** şablonu bulunamadı "
            f"(`templates/{normalize_tr(firma)}_{'alis' if islem_turu=='Alış' else 'satis'}.xlsx`). "
            "Yine de PDF üretilecek.",
            icon="⚠️",
        )

    run_disabled = file_bytes is None
    if st.button("🚀 Çalıştır", type="primary", disabled=run_disabled, use_container_width=True):
        progress_bar = st.progress(0)
        status_text = st.empty()

        try:
            zip_bytes = run_job(
                firma=firma,
                islem_turu=islem_turu,
                input_bytes=file_bytes,
                col_map=col_map,
                group_mode=group_mode,
                progress_bar=progress_bar,
                status_text=status_text,
            )

            st.download_button(
                label="⬇️ ZIP Dosyasını İndir",
                data=zip_bytes,
                file_name="teslim_tesellum_ciktilari.zip",
                mime="application/zip",
                use_container_width=True,
                type="secondary",
            )

        except Exception as e:
            st.error(f"❌ Hata: {e}")


if __name__ == "__main__":
    main()
